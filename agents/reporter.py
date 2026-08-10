# -*- coding: utf-8 -*-
"""Reporter — экспорт лидов в CSV.

В таблицу попадают ТОЛЬКО квалифицированные лиды: бизнесу нужен сайт и у него
есть разрешённый текущим режимом канал связи.

Экспорт включает основные поля лида и детерминированные resolver/audit evidence,
но не включает HTML, provider payloads или секреты.

Файл пишется в UTF-8 с BOM (Excel сразу показывает кириллицу правильно),
разделитель ";" (Excel с украинской/русской локалью корректно делит колонки).
Сначала идут бизнесы без сайта, затем с плохим сайтом.
"""

import csv
from datetime import datetime
from pathlib import Path
from typing import List

import config
from contactability import normalized_instagram_profile
from models import Business

# Колонки CSV: (заголовок, как достать значение из Business)
COLUMNS = [
    ("Business Name", lambda b: b.name),
    ("City", lambda b: b.city),
    ("Phone", lambda b: b.contactability.normalized_phone or ""),
    ("Email", lambda b: b.contactability.normalized_email or ""),
    ("Instagram URL", lambda b: b.instagram_url),
    (
        "Preferred Contact",
        lambda b: (
            b.preferred_contact_channel.value
            if b.preferred_contact_channel is not None
            else ""
        ),
    ),
    (
        "Contact Channels",
        lambda b: ",".join(channel.value for channel in b.contactability.channels),
    ),
    ("Website URL", lambda b: b.effective_website_url),
    ("Website Status", lambda b: b.website_status),
    ("Resolution Status", lambda b: b.website_resolution_status),
    ("Resolution Source", lambda b: b.website_resolution_source),
    ("Resolution Confidence", lambda b: f"{b.website_resolution_confidence:.3f}"),
    ("Resolution Evidence", lambda b: b.website_resolution_evidence),
    ("Resolution Error", lambda b: b.website_resolution_error),
    ("Lead Decision", lambda b: b.lead_decision),
    ("Lead Decision Reason", lambda b: b.lead_decision_reason),
]


def export_csv(businesses: List[Business], task_id: int = 0) -> str:
    """Записать CSV только с качественными лидами, вернуть путь к файлу."""
    config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = config.EXPORT_DIR / f"leads_task{task_id}_{stamp}.csv"

    # Страховка: даже если на вход пришли все бизнесы — пишем только лидов.
    leads = [b for b in businesses if b.is_lead]
    # Сначала "no website" (главный приоритет), потом "bad website".
    leads.sort(key=lambda b: 0 if b.website_status == "no website" else 1)

    # utf-8-sig = UTF-8 с BOM; newline="" обязателен для модуля csv
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([header for header, _ in COLUMNS])
        for b in leads:
            writer.writerow([str(getter(b) or "") for _, getter in COLUMNS])

    return str(path)


# Ширина колонок Excel (в символах) — чтобы всё было нормально видно
EXCEL_WIDTHS = {
    "A": 42,
    "B": 20,
    "C": 22,
    "D": 34,
    "E": 40,
    "F": 20,
    "G": 30,
    "H": 45,
    "I": 20,
    "J": 22,
    "K": 20,
    "L": 22,
    "M": 60,
    "N": 36,
    "O": 18,
    "P": 34,
}


def export_excel(businesses: List[Business], task_id: int = 0) -> str:
    """Записать .xlsx только с лидами и нормальной шириной колонок.

    Возвращает путь к файлу или "" если openpyxl недоступен (тогда работаем
    только с CSV). Колонки те же, что в CSV.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return ""

    config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = config.EXPORT_DIR / f"leads_task{task_id}_{stamp}.xlsx"

    leads = [b for b in businesses if b.is_lead]
    leads.sort(key=lambda b: 0 if b.website_status == "no website" else 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Заголовок
    ws.append([header for header, _ in COLUMNS])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Данные
    for b in leads:
        ws.append([str(getter(b) or "") for _, getter in COLUMNS])

    # Ширина колонок + закрепление шапки
    for col, width in EXCEL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    wb.save(path)
    return str(path)


def format_leads_summary(businesses: List[Business], limit: int = 20) -> str:
    """Format qualified leads with at least one usable contact per entry."""

    leads = [business for business in businesses if business.is_lead]
    leads.sort(key=lambda business: 0 if business.website_status == "no website" else 1)
    if not leads:
        ending = (
            "усі або з нормальним сайтом, або без доступного контакту"
            if config.LEAD_CONTACTABILITY_MODE == "multi_channel"
            else "усі або з нормальним сайтом, або без Instagram"
        )
        return f"Якісних лідів не знайдено ({ending})."

    blocks = []
    for business in leads[:limit]:
        contactability = business.contactability
        contact_lines = []
        if contactability.instagram_available:
            contact_lines.append(
                f"Instagram: {normalized_instagram_profile(business.instagram_url)}"
            )
        elif (
            config.LEAD_CONTACTABILITY_MODE == "instagram_only"
            and business.instagram_url
        ):
            contact_lines.append(f"Instagram: {business.instagram_url}")
        if contactability.normalized_phone:
            contact_lines.append(f"Телефон: {contactability.normalized_phone}")
        if contactability.normalized_email:
            contact_lines.append(f"Email: {contactability.normalized_email}")
        website_line = (
            f"\nWebsite: {business.effective_website_url}"
            if business.effective_website_url
            else ""
        )
        blocks.append(
            f"Назва: {business.name}\n"
            f"Місто: {business.city}\n"
            f"{chr(10).join(contact_lines)}\n"
            f"Статус сайту: {business.website_status}"
            f"{website_line}"
        )
    text = "\n\n".join(blocks)
    if len(leads) > limit:
        text += f"\n\n…та ще {len(leads) - limit} лідів — дивись CSV."
    return text


if __name__ == "__main__":
    # Ручной запуск на тестовых данных
    import json
    import sys

    sys.stdout.reconfigure(encoding="utf-8")
    data = json.loads(
        (Path(__file__).parent.parent / "tests" / "stage5_result.json").read_text(encoding="utf-8")
    )
    items = [Business(**d) for d in data]
    csv_path = export_csv(items, task_id=0)
    print(f"CSV создан: {csv_path}")
