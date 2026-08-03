import csv
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from agents import reporter
from models import Business


HEADERS = [
    "Business Name", "City", "Instagram URL", "Website URL", "Website Status",
    "Resolution Status", "Resolution Source", "Resolution Confidence",
    "Resolution Evidence", "Resolution Error", "Lead Decision", "Lead Decision Reason",
]


def lead():
    return Business(
        name="Alpha",
        city="Kyiv",
        instagram_url="https://instagram.com/alpha",
        website="https://maps.example/",
        website_resolved_url="https://official.example/",
        site_quality="bad",
        website_resolution_status="found_official",
        website_resolution_source="web_search",
        website_resolution_confidence=0.75,
        website_resolution_evidence='[{"schema_version":1}]',
        lead_decision="lead",
        lead_decision_reason="official_site_bad",
    )


class ReporterEvidenceTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.patch = patch.object(reporter.config, "EXPORT_DIR", Path(self.temp.name))
        self.patch.start()

    def tearDown(self):
        self.patch.stop()
        self.temp.cleanup()

    def test_columns_and_csv_values(self):
        excluded = Business(
            name="Uncertain", instagram_url="https://instagram.com/u",
            site_quality="technical_error",
        )
        path = reporter.export_csv([lead(), excluded], task_id=3)
        with open(path, encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=";"))
        self.assertEqual(rows[0], HEADERS)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][3], "https://official.example/")
        self.assertEqual(rows[1][7], "0.750")
        self.assertEqual(rows[1][8], '[{"schema_version":1}]')

    def test_excel_matches_csv_contract(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        path = reporter.export_excel([lead()], task_id=4)
        workbook = load_workbook(path, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        self.assertEqual(list(rows[0]), HEADERS)
        self.assertEqual(rows[1][3], "https://official.example/")
        self.assertEqual(rows[1][7], "0.750")

    def test_telegram_is_concise_and_includes_effective_url(self):
        text = reporter.format_leads_summary([lead()])
        self.assertIn("Website: https://official.example/", text)
        self.assertNotIn("schema_version", text)


if __name__ == "__main__":
    unittest.main()
